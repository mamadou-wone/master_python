# import openpyxl
# from datetime import datetime

# # Load the Excel file
# file_path = '2020.xlsx'
# workbook = openpyxl.load_workbook(file_path)

# # Define the columns you want to print
# columns_to_print = ['A', 'B', 'D', 'G', 'J', 'O', 'R']

# # Function to format the date
# def format_date(date):
#     return date.strftime('%m/%d/%Y')

# # Function to print row data from a sheet
# def print_sheet(sheet):
#     # Start from row 11 and print values from specified columns
#     for row in range(11, sheet.max_row + 1):
#         row_data = []
#         for column in columns_to_print:
#             cell_value = sheet[column + str(row)].value
#             if isinstance(cell_value, datetime):
#                 cell_value = format_date(cell_value)
#             row_data.append(cell_value)
#         if all(value is None for value in row_data):
#             break  # Exit the loop if all values in the row are None
#         # Add values of G4 and H4 at the end
#         row_data.append(sheet['G4'].value)
#         row_data.append(sheet['H4'].value)
#         print(row_data)

# # Iterate over all sheets and print their contents
# for sheet_name in workbook.sheetnames:
#     print(f"Sheet: {sheet_name}")
#     sheet = workbook[sheet_name]
#     print_sheet(sheet)
#     print()  # Print an empty line between sheets

# # Close the workbook
# workbook.close()

import openpyxl
import csv
from datetime import datetime

# Load the Excel file
file_path = 'JOURNALAS2024.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Define the columns you want to print
# columns_to_print = ['A', 'B', 'D', 'G', 'J', 'O', 'R']
columns_to_print = ['A', 'B', 'C', 'E', 'G', 'J', 'L']

# Function to format the date
def format_date(date):
    return date.strftime('%d/%m/%Y')

# Function to print row data from a sheet
def write_sheet_to_csv(sheet, csv_writer):
    # Start from row 11 and print values from specified columns
    for row in range(11, sheet.max_row + 1):
        row_data = []
        for column in columns_to_print:
            cell_value = sheet[column + str(row)].value
            if isinstance(cell_value, datetime):
                cell_value = format_date(cell_value)
            row_data.append(cell_value)
        if all(value is None for value in row_data):
            break  # Exit the loop if all values in the row are None
        # Add values of G4 and H4 at the end
        # row_data.append(sheet['G4'].value)
        # row_data.append(sheet['H4'].value)
        row_data.append(sheet['E4'].value)
        row_data.append(sheet['F4'].value)
        csv_writer.writerow(row_data)

# Iterate over all sheets and write their contents to a CSV file
with open('JOURNALAS2024_treated__.csv', 'w', newline='') as csvfile:
    csv_writer = csv.writer(csvfile)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        write_sheet_to_csv(sheet, csv_writer)

# Close the workbook
workbook.close()
